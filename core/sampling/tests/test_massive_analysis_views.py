"""Pruebas para la vista de listado de análisis masivos (Metales Pesados)."""

from datetime import timedelta
from io import BytesIO

from django.contrib.auth.models import Permission
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from core.analytical_method.models import AnalyticalMethod, HeavyMetal
from core.laboratory.models import Laboratory
from core.product.models import AnalyticalMethodProduct, SpecificationProduct
from core.sampling.models import MassiveSampleAnalysis, SamplingAnalysis, SamplingProcess
from core.sampling.tests.factories import build_sample_point
from core.user.models import User


class MassiveSampleAnalysisListViewTests(TestCase):
    """Pruebas para MassiveSampleAnalysisListView (server-side DataTables y filtros)."""

    @classmethod
    def setUpTestData(cls):
        cls.point = build_sample_point(code='MP1')
        cls.site = cls.point.product.site
        cls.laboratory = Laboratory.objects.create(
            laboratory_name='Lab Test', site=cls.site
        )
        cls.user = User.objects.create_user(
            username='analista', password='test1234', laboratory=cls.laboratory,
            first_name='Ana', last_name='Lista'
        )
        cls.user.user_permissions.add(
            Permission.objects.get(codename='add_reagent')
        )
        cls.method = AnalyticalMethod.objects.create(
            description_analytical_method='Metales Pesados',
            code_analytical_method='MET-01',
            sample_size=1.0,
            type_method='Lectura Directa',
            laboratory=cls.laboratory,
        )
        cls.other_method = AnalyticalMethod.objects.create(
            description_analytical_method='Plomo',
            code_analytical_method='MET-02',
            sample_size=1.0,
            type_method='Lectura Directa',
            laboratory=cls.laboratory,
        )
        HeavyMetal.objects.create(
            analytical_method=cls.method, metal_description='Plomo', unit_measure='mg/L'
        )
        HeavyMetal.objects.create(
            analytical_method=cls.method, metal_description='Cadmio', unit_measure='mg/L'
        )
        cls.sampling = SamplingProcess.objects.create(
            point_sampling=cls.point,
            type_sampling='En Proceso',
            date_sampling_scheduled=timezone.now(),
            number_sample='MP1-20260101-1',
        )
        cls.analysis = MassiveSampleAnalysis.objects.create(
            sampling_process=cls.sampling,
            analytical_method=cls.method,
            result=0.1234,
            standard_deviation=0.01,
            coefficient_variation=0.5,
            comply='Cumple',
            date_analysis=timezone.now(),
            analized_by=cls.user,
        )
        cls.url = reverse('sampling:list_massive_sample_analysis')

    def setUp(self):
        self.client.force_login(self.user)

    def test_get_lista_renderiza_filtros(self):
        """La página carga con el filtro avanzado y la tabla."""
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id_start_date')
        self.assertContains(response, 'id_end_date')
        self.assertContains(response, 'id_analized_by')
        self.assertContains(response, 'id_analytical_method')
        self.assertContains(response, 'id_sample')

    def test_searchdata_pagina_y_excluye_campos(self):
        """searchdata retorna JSON paginado sin desviación, coeficiente ni concepto."""
        response = self.client.post(self.url, {
            'action': 'searchdata', 'draw': 1, 'start': 0, 'length': 10,
        })
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['recordsTotal'], 1)
        self.assertEqual(payload['recordsFiltered'], 1)
        row = payload['data'][0]
        self.assertNotIn('standard_deviation', row)
        self.assertNotIn('coefficient_variation', row)
        self.assertNotIn('comply', row)
        self.assertEqual(row['sampling_process'], 'MP1-20260101-1')
        self.assertEqual(row['product'], 'Producto Test')
        self.assertEqual(row['analytical_method'], 'Metales Pesados')
        self.assertEqual(row['metal'], 'Plomo, Cadmio')
        self.assertEqual(row['result'], '0.1234')

    def test_filtro_por_rango_de_fechas(self):
        """El filtro de fechas excluye registros fuera del rango."""
        yesterday = (timezone.now() - timedelta(days=1)).date().isoformat()
        tomorrow = (timezone.now() + timedelta(days=1)).date().isoformat()

        dentro = self.client.post(self.url, {
            'action': 'searchdata', 'draw': 1, 'start': 0, 'length': 10,
            'start_date': yesterday, 'end_date': tomorrow,
        }).json()
        self.assertEqual(dentro['recordsFiltered'], 1)

        fuera = self.client.post(self.url, {
            'action': 'searchdata', 'draw': 1, 'start': 0, 'length': 10,
            'start_date': tomorrow, 'end_date': tomorrow,
        }).json()
        self.assertEqual(fuera['recordsFiltered'], 0)

    def test_filtro_por_analista_metodo_y_muestra(self):
        """Los filtros de analista, método y muestra reducen los resultados."""
        base = {'action': 'searchdata', 'draw': 1, 'start': 0, 'length': 10}

        por_metodo = self.client.post(self.url, {**base, 'analytical_method': str(self.other_method.id)}).json()
        self.assertEqual(por_metodo['recordsFiltered'], 0)

        por_muestra = self.client.post(self.url, {**base, 'sample': 'MP1-20260101'}).json()
        self.assertEqual(por_muestra['recordsFiltered'], 1)

        por_muestra_inexistente = self.client.post(self.url, {**base, 'sample': 'NO-EXISTE'}).json()
        self.assertEqual(por_muestra_inexistente['recordsFiltered'], 0)

        otro_usuario = User.objects.create_user(username='otro', password='x1234567')
        por_analista = self.client.post(self.url, {**base, 'analized_by': str(otro_usuario.id)}).json()
        self.assertEqual(por_analista['recordsFiltered'], 0)

    def test_sin_laboratorio_no_ve_registros(self):
        """Un usuario sin laboratorio asignado no ve análisis."""
        sin_lab = User.objects.create_user(username='sinlab', password='x1234567')
        sin_lab.user_permissions.add(Permission.objects.get(codename='add_reagent'))
        self.client.force_login(sin_lab)
        payload = self.client.post(self.url, {
            'action': 'searchdata', 'draw': 1, 'start': 0, 'length': 10,
        }).json()
        self.assertEqual(payload['recordsTotal'], 0)


def build_excel_upload(rows, headers=None):
    """Construye un archivo .xlsx en memoria con los encabezados y filas dados."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers or [
        'Muestra', 'Método de Análisis', 'Fecha de Análisis', 'Realizado por', 'Plomo', 'Cadmio',
    ])
    for row in rows:
        sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    return SimpleUploadedFile(
        'cargue.xlsx', buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


class MassiveSampleAnalysisUploadViewTests(TestCase):
    """Pruebas para MassiveSampleAnalysisUploadView (cargue masivo desde Excel)."""

    @classmethod
    def setUpTestData(cls):
        cls.point = build_sample_point(code='MP2')
        cls.site = cls.point.product.site
        cls.laboratory = Laboratory.objects.create(
            laboratory_name='Lab Upload', site=cls.site
        )
        cls.user = User.objects.create_user(
            username='cargador', password='test1234', laboratory=cls.laboratory,
            first_name='Carlos', last_name='Cargas'
        )
        cls.user.user_permissions.add(
            Permission.objects.get(codename='add_reagent')
        )
        cls.method = AnalyticalMethod.objects.create(
            description_analytical_method='Metales Pesados',
            code_analytical_method='MET-01',
            sample_size=1.0,
            type_method='Lectura Directa',
            laboratory=cls.laboratory,
        )
        cls.lead = HeavyMetal.objects.create(
            analytical_method=cls.method, metal_description='Plomo',
            unit_measure='mg/L', quantification_limit=0.05,
        )
        cls.cadmium = HeavyMetal.objects.create(
            analytical_method=cls.method, metal_description='Cadmio',
            unit_measure='mg/L', quantification_limit=0.01,
        )
        cls.sampling = SamplingProcess.objects.create(
            point_sampling=cls.point,
            type_sampling='En Proceso',
            date_sampling_scheduled=timezone.now(),
            number_sample='MP2-20260101-1',
        )
        cls.url = reverse('sampling:upload_massive_sample_analysis')

    def setUp(self):
        self.client.force_login(self.user)

    def test_get_renderiza_formulario_modal(self):
        """GET retorna el formulario de cargue para el modal."""
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'upload_form')
        self.assertContains(response, 'upload_progress_bar')

    def test_cargue_crea_un_registro_por_metal(self):
        """Una fila con dos metales crea dos registros con su metal asociado."""
        excel = build_excel_upload([[
            'MP2-20260101-1', 'Metales Pesados', '2026-01-15 10:30', 'cargador', 1.5, 0.25,
        ]])
        response = self.client.post(self.url, {'action': 'upload', 'file': excel})
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['created'], 2)
        self.assertEqual(payload['total_rows'], 1)
        self.assertEqual(payload['errors'], [])

        registros = MassiveSampleAnalysis.objects.filter(sampling_process=self.sampling)
        self.assertEqual(registros.count(), 2)
        por_metal = {r.heavy_metal.metal_description: r for r in registros}
        self.assertEqual(por_metal['Plomo'].result, 1.5)
        self.assertEqual(por_metal['Cadmio'].result, 0.25)
        self.assertEqual(por_metal['Plomo'].analized_by, self.user)
        self.assertEqual(por_metal['Plomo'].analytical_method, self.method)

    def test_resultado_cero_o_negativo_usa_limite_cuantificacion(self):
        """Resultados <= 0 se reemplazan por el límite de cuantificación del metal."""
        excel = build_excel_upload([[
            'MP2-20260101-1', 'MET-01', '2026-01-15', 'cargador', 0, -3.2,
        ]])
        payload = self.client.post(self.url, {'action': 'upload', 'file': excel}).json()
        self.assertEqual(payload['created'], 2)
        por_metal = {r.heavy_metal.metal_description: r for r in
                     MassiveSampleAnalysis.objects.filter(sampling_process=self.sampling)}
        self.assertEqual(por_metal['Plomo'].result, 0.05)
        self.assertEqual(por_metal['Cadmio'].result, 0.01)

    def test_fila_con_datos_invalidos_reporta_errores(self):
        """Filas con muestra, método o analista inexistentes no se crean y se reportan."""
        excel = build_excel_upload([
            ['NO-EXISTE', 'Metales Pesados', '2026-01-15', 'cargador', 1.0, 1.0],
            ['MP2-20260101-1', 'Método Falso', '2026-01-15', 'cargador', 1.0, 1.0],
            ['MP2-20260101-1', 'Metales Pesados', '2026-01-15', 'fantasma', 1.0, 1.0],
        ])
        payload = self.client.post(self.url, {'action': 'upload', 'file': excel}).json()
        self.assertEqual(payload['created'], 0)
        self.assertEqual(len(payload['errors']), 3)
        self.assertFalse(MassiveSampleAnalysis.objects.exists())

    def test_estructura_invalida_retorna_error(self):
        """Un Excel sin las columnas fijas esperadas retorna error."""
        excel = build_excel_upload(
            [['MP2-20260101-1', 'Metales Pesados', '2026-01-15', 'cargador', 1.0]],
            headers=['Sample', 'Método de Análisis', 'Fecha de Análisis', 'Realizado por', 'Plomo'],
        )
        payload = self.client.post(self.url, {'action': 'upload', 'file': excel}).json()
        self.assertIn('error', payload)
        self.assertFalse(MassiveSampleAnalysis.objects.exists())

    def test_fecha_como_serial_numerico_de_excel(self):
        """Una fecha guardada como número de serie de Excel (formato numérico) se convierte."""
        from datetime import date as date_class
        serial = (date_class(2026, 1, 15) - date_class(1899, 12, 30)).days
        excel = build_excel_upload([[
            'MP2-20260101-1', 'Metales Pesados', float(serial), 'cargador', 1.5,
        ]], headers=['Muestra', 'Método de Análisis', 'Fecha de Análisis', 'Realizado por', 'Plomo'])
        payload = self.client.post(self.url, {'action': 'upload', 'file': excel}).json()
        self.assertEqual(payload['created'], 1)
        registro = MassiveSampleAnalysis.objects.get()
        self.assertEqual(registro.date_analysis.date().isoformat(), '2026-01-15')

    def test_fecha_serial_invalida_reporta_error_de_formato(self):
        """Un serial fuera de rango reporta error de formato, no de obligatoriedad."""
        excel = build_excel_upload([[
            'MP2-20260101-1', 'Metales Pesados', 999999999999, 'cargador', 1.5,
        ]], headers=['Muestra', 'Método de Análisis', 'Fecha de Análisis', 'Realizado por', 'Plomo'])
        payload = self.client.post(self.url, {'action': 'upload', 'file': excel}).json()
        self.assertEqual(payload['created'], 0)
        self.assertIn('formato válido', payload['errors'][0])

    def test_fecha_vacia_reporta_obligatoriedad(self):
        """Una celda de fecha vacía reporta que la fecha es obligatoria."""
        excel = build_excel_upload([[
            'MP2-20260101-1', 'Metales Pesados', None, 'cargador', 1.5,
        ]], headers=['Muestra', 'Método de Análisis', 'Fecha de Análisis', 'Realizado por', 'Plomo'])
        payload = self.client.post(self.url, {'action': 'upload', 'file': excel}).json()
        self.assertEqual(payload['created'], 0)
        self.assertIn('obligatoria', payload['errors'][0])

    def test_archivo_no_xlsx_es_rechazado(self):
        """Un archivo con extensión diferente a .xlsx es rechazado por el formulario."""
        archivo = SimpleUploadedFile('cargue.csv', b'a,b,c', content_type='text/csv')
        payload = self.client.post(self.url, {'action': 'upload', 'file': archivo}).json()
        self.assertIn('error', payload)

    def test_cargue_soporta_1000_filas(self):
        """El cargue procesa al menos 1000 filas en una sola petición."""
        filas = [['MP2-20260101-1', 'Metales Pesados', '2026-01-15 10:30', 'cargador', 1.5]
                 for _ in range(1000)]
        excel = build_excel_upload(filas, headers=[
            'Muestra', 'Método de Análisis', 'Fecha de Análisis', 'Realizado por', 'Plomo',
        ])
        payload = self.client.post(self.url, {'action': 'upload', 'file': excel}).json()
        self.assertEqual(payload['total_rows'], 1000)
        self.assertEqual(payload['created'], 1000)
        self.assertEqual(payload['errors'], [])
        self.assertEqual(MassiveSampleAnalysis.objects.count(), 1000)

    def test_sin_permiso_redirige(self):
        """Un usuario sin permiso de cargue es redirigido."""
        sin_permiso = User.objects.create_user(
            username='sinpermiso', password='x1234567', laboratory=self.laboratory
        )
        self.client.force_login(sin_permiso)
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 302)

    def _upload(self, rows, headers=None):
        """Carga un Excel con las filas dadas y retorna el payload JSON."""
        excel = build_excel_upload(rows, headers=headers)
        return self.client.post(self.url, {'action': 'upload', 'file': excel}).json()

    def _build_spec(self, lower, upper):
        """Crea la especificación del producto asociada al método de análisis."""
        method_product = AnalyticalMethodProduct.objects.create(
            product=self.point.product, analytical_method=self.method
        )
        return SpecificationProduct.objects.create(
            product=self.point.product, type_test='Químico', test_prod='Metales Pesados',
            method_test=method_product, lower_limit_prod=lower, upper_limit_prod=upper,
        )

    def test_suma_resultados_crea_sampling_analysis(self):
        """La suma de resultados se asigna al average_concentration del SamplingAnalysis."""
        payload = self._upload([[
            'MP2-20260101-1', 'Metales Pesados', '2026-01-15 10:30', 'cargador', 1.5, 0.25,
        ]])
        self.assertEqual(payload['created'], 2)

        analysis = SamplingAnalysis.objects.get(
            sampling_process=self.sampling, analytical_method=self.method
        )
        self.assertAlmostEqual(analysis.average_concentration, 1.75)
        # Sin especificación del producto el concepto queda vacío
        self.assertIsNone(analysis.comply)

    def test_suma_resultados_actualiza_sampling_analysis_existente(self):
        """Un SamplingAnalysis existente se actualiza con la nueva suma de resultados."""
        existente = SamplingAnalysis.objects.create(
            sampling_process=self.sampling, analytical_method=self.method,
            average_concentration=9.9,
        )
        self._upload([[
            'MP2-20260101-1', 'Metales Pesados', '2026-01-15 10:30', 'cargador', 1.0, 1.0,
        ]])
        existente.refresh_from_db()
        self.assertAlmostEqual(existente.average_concentration, 2.0)
        self.assertEqual(SamplingAnalysis.objects.count(), 1)

    def test_resultado_negativo_suma_limite_cuantificacion(self):
        """Un resultado negativo aporta el límite de cuantificación a la suma."""
        self._upload([[
            'MP2-20260101-1', 'Metales Pesados', '2026-01-15', 'cargador', -1.0, 0.25,
        ]])
        analysis = SamplingAnalysis.objects.get(
            sampling_process=self.sampling, analytical_method=self.method
        )
        self.assertAlmostEqual(analysis.average_concentration, 0.05 + 0.25)

    def test_comply_cumple_dentro_de_limites_inclusive(self):
        """Comply es 'Cumple' cuando la suma está dentro o igual a los límites."""
        self._build_spec(lower=1.0, upper=2.0)

        self._upload([[
            'MP2-20260101-1', 'Metales Pesados', '2026-01-15', 'cargador', 1.5, 0.25,
        ]])
        analysis = SamplingAnalysis.objects.get(
            sampling_process=self.sampling, analytical_method=self.method
        )
        self.assertAlmostEqual(analysis.average_concentration, 1.75)
        self.assertEqual(analysis.comply, 'Cumple')

    def test_comply_cumple_en_limite_exacto(self):
        """Comply es 'Cumple' cuando la suma es igual al límite superior."""
        self._build_spec(lower=1.0, upper=2.0)
        self._upload([[
            'MP2-20260101-1', 'Metales Pesados', '2026-01-15', 'cargador', 1.0, 1.0,
        ]])
        analysis = SamplingAnalysis.objects.get(
            sampling_process=self.sampling, analytical_method=self.method
        )
        self.assertAlmostEqual(analysis.average_concentration, 2.0)
        self.assertEqual(analysis.comply, 'Cumple')

    def test_comply_no_cumple_fuera_de_rango(self):
        """Comply es 'No Cumple' cuando la suma supera el límite superior."""
        self._build_spec(lower=1.0, upper=2.0)
        self._upload([[
            'MP2-20260101-1', 'Metales Pesados', '2026-01-15', 'cargador', 2.0, 1.5,
        ]])
        analysis = SamplingAnalysis.objects.get(
            sampling_process=self.sampling, analytical_method=self.method
        )
        self.assertAlmostEqual(analysis.average_concentration, 3.5)
        self.assertEqual(analysis.comply, 'No Cumple')

    def test_comply_no_cumple_bajo_limite_inferior(self):
        """Comply es 'No Cumple' cuando la suma está por debajo del límite inferior."""
        self._build_spec(lower=1.0, upper=5.0)
        self._upload([[
            'MP2-20260101-1', 'Metales Pesados', '2026-01-15', 'cargador', 0.3, 0.2,
        ]])
        analysis = SamplingAnalysis.objects.get(
            sampling_process=self.sampling, analytical_method=self.method
        )
        self.assertEqual(analysis.comply, 'No Cumple')

    def test_comply_vacio_con_limites_nulos(self):
        """Comply queda vacío cuando la especificación tiene los límites nulos."""
        self._build_spec(lower=None, upper=None)
        self._upload([[
            'MP2-20260101-1', 'Metales Pesados', '2026-01-15', 'cargador', 1.0, 1.0,
        ]])
        analysis = SamplingAnalysis.objects.get(
            sampling_process=self.sampling, analytical_method=self.method
        )
        self.assertIsNone(analysis.comply)

    def test_average_concentration_usa_cifras_significativas_del_metodo(self):
        """La suma se guarda redondeada a las cifras significativas del método (sig_figs_result)."""
        self.method.sig_figs_result = 3
        self.method.save()
        self._upload([[
            'MP2-20260101-1', 'Metales Pesados', '2026-01-15', 'cargador', 1.23456, 0.00001,
        ]])
        analysis = SamplingAnalysis.objects.get(
            sampling_process=self.sampling, analytical_method=self.method
        )
        self.assertEqual(analysis.average_concentration, round(1.23456 + 0.00001, 3))

    def test_average_concentration_cifras_significativas_por_defecto(self):
        """Con sig_figs_result por defecto (2) la suma se redondea a 2 decimales."""
        self._upload([[
            'MP2-20260101-1', 'Metales Pesados', '2026-01-15', 'cargador', 1.257, None,
        ]])
        analysis = SamplingAnalysis.objects.get(
            sampling_process=self.sampling, analytical_method=self.method
        )
        self.assertEqual(analysis.average_concentration, round(1.257, 2))

    def test_resultado_negativo_sin_limite_configurado_reporta_error(self):
        """Un negativo de un metal sin límite de cuantificación no se carga y se reporta."""
        HeavyMetal.objects.create(
            analytical_method=self.method, metal_description='Zinc', unit_measure='mg/L'
        )
        payload = self._upload(
            [['MP2-20260101-1', 'Metales Pesados', '2026-01-15', 'cargador', 1.5, -2.0]],
            headers=['Muestra', 'Método de Análisis', 'Fecha de Análisis', 'Realizado por',
                     'Plomo', 'Zinc'],
        )
        self.assertEqual(payload['created'], 1)  # Solo Plomo
        self.assertEqual(len(payload['errors']), 1)
        self.assertIn('Zinc', payload['errors'][0])
        self.assertIn('límite de cuantificación', payload['errors'][0])

        # El negativo no queda guardado ni afecta la suma
        self.assertFalse(
            MassiveSampleAnalysis.objects.filter(heavy_metal__metal_description='Zinc').exists()
        )
        analysis = SamplingAnalysis.objects.get(
            sampling_process=self.sampling, analytical_method=self.method
        )
        self.assertAlmostEqual(analysis.average_concentration, 1.5)
        self.assertGreaterEqual(analysis.average_concentration, 0)

    def test_resultado_cero_sin_limite_configurado_reporta_error(self):
        """Un cero de un metal sin límite de cuantificación tampoco se carga."""
        HeavyMetal.objects.create(
            analytical_method=self.method, metal_description='Zinc', unit_measure='mg/L'
        )
        payload = self._upload(
            [['MP2-20260101-1', 'Metales Pesados', '2026-01-15', 'cargador', 0]],
            headers=['Muestra', 'Método de Análisis', 'Fecha de Análisis', 'Realizado por', 'Zinc'],
        )
        self.assertEqual(payload['created'], 0)
        self.assertEqual(len(payload['errors']), 1)
        self.assertFalse(MassiveSampleAnalysis.objects.exists())

    def test_comply_se_evalua_con_valor_redondeado(self):
        """El concepto se evalúa con el average_concentration ya redondeado."""
        self._build_spec(lower=1.0, upper=2.0)
        self._upload([[
            'MP2-20260101-1', 'Metales Pesados', '2026-01-15', 'cargador', 2.004, None,
        ]])
        analysis = SamplingAnalysis.objects.get(
            sampling_process=self.sampling, analytical_method=self.method
        )
        self.assertEqual(analysis.average_concentration, 2.0)
        self.assertEqual(analysis.comply, 'Cumple')


class MassiveSampleAnalysisTemplateViewTests(TestCase):
    """Pruebas para MassiveSampleAnalysisTemplateView (descarga de plantilla Excel)."""

    @classmethod
    def setUpTestData(cls):
        cls.point = build_sample_point(code='MP3')
        cls.site = cls.point.product.site
        cls.laboratory = Laboratory.objects.create(
            laboratory_name='Lab Template', site=cls.site
        )
        cls.user = User.objects.create_user(
            username='plantillero', password='test1234', laboratory=cls.laboratory,
        )
        cls.user.user_permissions.add(
            Permission.objects.get(codename='add_reagent')
        )
        cls.method = AnalyticalMethod.objects.create(
            description_analytical_method='Metales Pesados',
            code_analytical_method='MET-01',
            sample_size=1.0,
            type_method='Lectura Directa',
            laboratory=cls.laboratory,
        )
        HeavyMetal.objects.create(
            analytical_method=cls.method, metal_description='Plomo', unit_measure='mg/L'
        )
        HeavyMetal.objects.create(
            analytical_method=cls.method, metal_description='Cadmio', unit_measure='mg/L'
        )
        cls.url = reverse('sampling:download_massive_sample_analysis_template')

    def setUp(self):
        self.client.force_login(self.user)

    def _load_template(self, response):
        """Carga el contenido de la respuesta como libro de Excel."""
        return load_workbook(BytesIO(response.content))

    def test_descarga_archivo_excel(self):
        """GET retorna un .xlsx adjunto con el nombre y tipo de contenido correctos."""
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn('Plantilla_Metales_Pesados.xlsx', response['Content-Disposition'])

    def test_encabezados_fijos_y_metales_ordenados(self):
        """La hoja Datos tiene las 4 columnas fijas y una columna por metal."""
        workbook = self._load_template(self.client.get(self.url))
        sheet = workbook['Datos']
        headers = [cell.value for cell in sheet[1]]
        self.assertEqual(
            headers,
            ['Muestra', 'Método de Análisis', 'Fecha de Análisis', 'Realizado por', 'Cadmio', 'Plomo'],
        )
        # Solo encabezados: lista para pegar o digitar registros
        self.assertEqual(sheet.max_row, 1)

    def test_hoja_instrucciones_documenta_metales_por_metodo(self):
        """La hoja Instrucciones existe y lista los metales del método."""
        workbook = self._load_template(self.client.get(self.url))
        self.assertIn('Instrucciones', workbook.sheetnames)
        contenido = '\n'.join(
            str(row[0]) for row in workbook['Instrucciones'].iter_rows(values_only=True)
        )
        self.assertIn('MET-01', contenido)
        self.assertIn('Plomo', contenido)
        self.assertIn('Cadmio', contenido)

    def test_sin_laboratorio_plantilla_solo_con_columnas_fijas(self):
        """Un usuario sin laboratorio descarga la plantilla solo con las columnas fijas."""
        sin_lab = User.objects.create_user(username='sinlab2', password='x1234567')
        sin_lab.user_permissions.add(
            Permission.objects.get(codename='add_reagent')
        )
        self.client.force_login(sin_lab)
        workbook = self._load_template(self.client.get(self.url))
        headers = [cell.value for cell in workbook['Datos'][1]]
        self.assertEqual(
            headers, ['Muestra', 'Método de Análisis', 'Fecha de Análisis', 'Realizado por']
        )

    def test_sin_permiso_redirige(self):
        """Un usuario sin permiso de consulta es redirigido."""
        sin_permiso = User.objects.create_user(
            username='sinpermiso2', password='x1234567', laboratory=self.laboratory
        )
        self.client.force_login(sin_permiso)
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 302)
