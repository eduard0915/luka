"""Lógica de negocio para la generación automática de muestras."""

import unicodedata
from datetime import date, datetime, time, timedelta
from io import BytesIO

from django.db import IntegrityError, transaction
from django.db.models import Q, Value
from django.db.models.functions import Concat
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel

from core.analytical_method.models import AnalyticalMethod, HeavyMetal
from core.product.models import SpecificationProduct
from core.sampling.models import (
    MassiveSampleAnalysis, SamplingAnalysis, SamplingGenerationLog, SamplingProcess,
    next_sample_number,
)
from core.user.models import User

DAILY_PERIODICITY = {'Diaria', 'Diario'}


def compute_sampling_times(group, target_date):
    """Calcula los horarios de muestreo espaciados por sample_frequency horas desde first_hour_sampling."""
    point = group.sampling_point
    freq = point.sample_frequency
    if not freq or freq <= 0:
        return []

    first = timezone.make_aware(
        datetime.combine(target_date, group.first_hour_sampling),
        timezone.get_current_timezone(),
    )

    times = []
    current = first
    end_of_day = first + timedelta(days=1)

    while current < end_of_day:
        times.append(current)
        current += timedelta(hours=freq)

    return times


def should_skip_group(group):
    """Determina si un grupo de muestreo debe omitirse por no cumplir las condiciones."""
    point = group.sampling_point
    return (
        not group.enable_sampling_group
        or group.number_sampling_day < 1
        or not point.enable_point
        or not point.sample_point_code
        or point.periodicity not in DAILY_PERIODICITY
        or not point.sample_frequency
        or point.sample_frequency <= 0
    )


def generate_samplings_for_group(group, target_date):
    """Crea el lote de muestras de un grupo para un día. Retorna el log o None si ya existía."""
    try:
        with transaction.atomic():
            log = SamplingGenerationLog.objects.create(
                sampling_group=group,
                target_date=target_date,
                skipped=should_skip_group(group),
            )
            if log.skipped:
                return log
            point = group.sampling_point
            for scheduled_at in compute_sampling_times(group, target_date):
                SamplingProcess.objects.create(
                    group_sampling=group,
                    type_sampling='En Proceso',
                    date_sampling_scheduled=scheduled_at,
                    automatic_sampling=True,
                    number_sample=next_sample_number(point, target_date),
                )
                log.samples_created += 1
            log.save(update_fields=['samples_created'])
            return log
    except IntegrityError:
        # Solo el choque con unique_group_target_date significa 'ya generado'.
        # Cualquier otra violación (FK a un grupo borrado, NOT NULL, ...) es un error
        # real: debe escalar al comando, que la cuenta y termina con CommandError.
        if SamplingGenerationLog.objects.filter(
            sampling_group=group, target_date=target_date
        ).exists():
            return None
        raise


# ---------------------------------------------------------------------------
# Cargue masivo de análisis de metales pesados (MassiveSampleAnalysis)
# ---------------------------------------------------------------------------

MAX_ROW_ERRORS = 100
DATE_INPUT_FORMATS = (
    '%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d',
    '%d/%m/%Y %H:%M:%S', '%d/%m/%Y %H:%M', '%d/%m/%Y',
)
EXPECTED_FIXED_HEADERS = ('metal', 'metodo', 'realizado por', 'fecha')


def _normalize_header(value):
    """Normaliza un encabezado de Excel: minúsculas, sin tildes ni espacios extra."""
    text = unicodedata.normalize('NFKD', str(value or ''))
    return ''.join(c for c in text if not unicodedata.combining(c)).strip().lower()


def _parse_date_analysis(value):
    """Convierte una celda de fecha (datetime, date, serial de Excel o texto) a datetime con zona horaria."""
    if value is None or (isinstance(value, str) and not value.strip()):
        raise ValueError('la Fecha de Análisis es obligatoria')

    if isinstance(value, datetime):
        parsed = value
    elif isinstance(value, date):
        parsed = datetime.combine(value, time.min)
    elif isinstance(value, (int, float)) and not isinstance(value, bool):
        # La celda tiene formato numérico/general: Excel guarda la fecha como serial
        try:
            converted = from_excel(value)
        except Exception:
            raise ValueError(
                f'la fecha "{value}" no tiene un formato válido (aaaa-mm-dd o dd/mm/aaaa)'
            )
        parsed = converted if isinstance(converted, datetime) else datetime.combine(converted, time.min)
    elif isinstance(value, str):
        parsed = None
        for fmt in DATE_INPUT_FORMATS:
            try:
                parsed = datetime.strptime(value.strip(), fmt)
                break
            except ValueError:
                continue
        if parsed is None:
            raise ValueError(f'la fecha "{value}" no tiene un formato válido (aaaa-mm-dd o dd/mm/aaaa)')
    else:
        raise ValueError(
            f'la fecha "{value}" no tiene un formato válido (aaaa-mm-dd o dd/mm/aaaa)'
        )

    if timezone.is_naive(parsed):
        parsed = timezone.make_aware(parsed, timezone.get_current_timezone())
    return parsed


def _read_excel_rows(excel_file):
    """Lee el Excel y retorna las muestras de la fila 1 y las filas de metales.

    Fila 1: Metal, Metodo, Realizado por, Fecha y una columna por cada muestra
    (número de sampling_process) desde la columna E en adelante.
    Filas 2+: una fila por metal con sus resultados en cada columna de muestra.
    """
    try:
        workbook = load_workbook(excel_file, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f'El archivo no es un Excel válido (.xlsx): {exc}')
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers or len(headers) < 5:
            raise ValueError('El archivo no tiene la estructura esperada (faltan columnas).')

        fixed = tuple(_normalize_header(h) for h in headers[:4])
        if fixed != EXPECTED_FIXED_HEADERS:
            raise ValueError(
                'Las columnas fijas deben ser: Metal, Metodo, Realizado por, '
                'Fecha y una columna por muestra desde la columna E.'
            )

        sample_headers = [str(h).strip() if h is not None else '' for h in headers[4:]]

        raw_rows = []
        for row in rows:
            if row is None or all(cell is None or str(cell).strip() == '' for cell in row):
                continue
            raw_rows.append(row)
        return sample_headers, raw_rows
    finally:
        workbook.close()


def _build_lookup_maps(sample_headers, raw_rows, site, laboratory):
    """Precarga muestras, métodos, analistas y metales referenciados en el archivo."""
    sample_numbers = {s.strip() for s in sample_headers if s and s.strip()}
    method_names = {_normalize_header(r[1]) for r in raw_rows if len(r) > 1 and r[1] is not None}
    analyst_names = {str(r[2]).strip() for r in raw_rows if len(r) > 2 and r[2] is not None}

    samples = SamplingProcess.objects.filter(
        Q(group_sampling__sampling_point__product__site=site) |
        Q(point_sampling__product__site=site),
        number_sample__in=sample_numbers,
    )
    sample_map = {s.number_sample.strip().lower(): s for s in samples}

    methods = AnalyticalMethod.objects.filter(laboratory=laboratory, enable_analytical_method=True)
    method_map = {}
    for m in methods:
        method_map.setdefault(_normalize_header(m.code_analytical_method), m)
        method_map.setdefault(_normalize_header(m.description_analytical_method), m)
    # Solo conserva los métodos realmente referenciados en el archivo
    referenced = {method_map[name] for name in method_names if name in method_map}

    metals = HeavyMetal.objects.filter(analytical_method__in=referenced)
    metal_map = {(m.analytical_method_id, _normalize_header(m.metal_description)): m for m in metals}

    analysts = User.objects.annotate(
        full_name=Concat('first_name', Value(' '), 'last_name')
    ).filter(Q(username__in=analyst_names) | Q(full_name__in=analyst_names))
    analyst_map = {}
    for u in analysts:
        analyst_map[u.username.strip().lower()] = u
        full_name = f'{u.first_name} {u.last_name}'.strip().lower()
        if full_name:
            analyst_map.setdefault(full_name, u)

    return sample_map, method_map, metal_map, analyst_map


def process_massive_analysis_excel(excel_file, user):
    """Procesa el cargue masivo de análisis de metales pesados desde un Excel.

    Estructura del archivo (una columna por muestra, una fila por metal):
      Fila 1: Metal | Metodo | Realizado por | Fecha | Muestra1 | Muestra2 | ...
      Filas 2+: una fila por metal y una columna de resultado por cada muestra.

    El número de muestra (sampling_process) se escribe en la fila 1 a partir de
    la columna E. Cada celda (fila de metal, columna de muestra) con un valor
    crea un MassiveSampleAnalysis para esa muestra y ese metal.

    Si un resultado es 0 o negativo se asigna el límite de cuantificación
    del metal correspondiente. La suma de los resultados de cada muestra y
    método se asigna al campo average_concentration del SamplingAnalysis
    asociado (se crea si no existe) y su concepto (comply) se evalúa contra
    los límites de la SpecificationProduct del producto para ese método.
    Retorna un resumen con creados y errores.
    """
    if not user.laboratory:
        raise ValueError('El usuario no tiene un laboratorio asignado.')

    sample_headers, raw_rows = _read_excel_rows(excel_file)
    sample_map, method_map, metal_map, analyst_map = _build_lookup_maps(
        sample_headers, raw_rows, user.laboratory.site, user.laboratory
    )

    objects = []
    errors = []
    totals = {}
    skipped = 0

    for row_number, row in enumerate(raw_rows, start=2):
        if len(errors) >= MAX_ROW_ERRORS:
            errors.append(f'Se alcanzó el máximo de {MAX_ROW_ERRORS} errores reportados.')
            break

        # Las filas sin resultado en ninguna muestra se ignoran: son metales
        # pre-diligenciados de la plantilla que el usuario no midió.
        if not _row_has_results(row, len(sample_headers)):
            skipped += 1
            continue

        row_errors = []
        metal_desc = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ''

        method = method_map.get(_normalize_header(row[1])) if len(row) > 1 and row[1] is not None else None
        if method is None:
            method_ref = row[1] if len(row) > 1 else None
            row_errors.append(f'el método de análisis "{method_ref}" no existe o está inhabilitado')

        analyst = analyst_map.get(str(row[2]).strip().lower()) if len(row) > 2 and row[2] is not None else None
        if analyst is None:
            analyst_ref = row[2] if len(row) > 2 else None
            row_errors.append(f'el analista "{analyst_ref}" no existe')

        try:
            date_analysis = _parse_date_analysis(row[3] if len(row) > 3 else None)
        except ValueError as exc:
            date_analysis = None
            row_errors.append(str(exc))

        metal = None
        if method is not None and metal_desc:
            metal = metal_map.get((method.id, _normalize_header(metal_desc)))
            if metal is None:
                row_errors.append(
                    f'el metal "{metal_desc}" no está asociado '
                    f'al método "{method.description_analytical_method}"'
                )

        if row_errors:
            errors.append(f'Fila {row_number}: ' + '; '.join(row_errors))
            continue

        for col_index, sample_ref in enumerate(sample_headers, start=5):
            value = row[col_index - 1] if len(row) > col_index - 1 else None
            if value is None or str(value).strip() == '':
                continue

            col_letter = get_column_letter(col_index)
            if not sample_ref:
                errors.append(
                    f'Fila {row_number}, columna {col_letter}: hay un resultado pero la '
                    f'columna no tiene número de muestra en la fila 1'
                )
                continue

            sample = sample_map.get(sample_ref.strip().lower())
            if sample is None:
                errors.append(
                    f'Fila {row_number}, columna {col_letter}: '
                    f'la muestra "{sample_ref}" no existe en el sitio'
                )
                continue

            try:
                result = float(value)
            except (TypeError, ValueError):
                errors.append(
                    f'Fila {row_number}, columna {col_letter} '
                    f'({metal_desc}, muestra {sample_ref}): el resultado no es numérico ({value})'
                )
                continue

            if result <= 0:
                if metal.quantification_limit is None:
                    errors.append(
                        f'Fila {row_number}, columna {col_letter}: '
                        f'el resultado es {result} pero el metal "{metal_desc}" no tiene '
                        f'límite de cuantificación configurado'
                    )
                    continue
                result = metal.quantification_limit

            objects.append(MassiveSampleAnalysis(
                sampling_process=sample,
                analytical_method=method,
                heavy_metal=metal,
                result=result,
                date_analysis=date_analysis,
                analized_by=analyst,
                user_creation=user,
            ))
            key = (sample.id, method.id)
            totals.setdefault(key, {'sample': sample, 'method': method, 'total': 0.0})
            totals[key]['total'] += result

    if errors:
        # Cargue atómico: si hay filas o datos no válidos no se guarda nada.
        return {
            'created': 0,
            'total_rows': len(raw_rows),
            'skipped': skipped,
            'saved': False,
            'errors': errors,
        }

    with transaction.atomic():
        if objects:
            MassiveSampleAnalysis.objects.bulk_create(objects, batch_size=1000)
            _update_sampling_analysis_results(totals, user)
            _set_samples_en_proceso(totals)

    return {
        'created': len(objects),
        'total_rows': len(raw_rows),
        'skipped': skipped,
        'saved': True,
        'errors': errors,
    }


def _row_has_results(row, n_samples):
    """Indica si una fila tiene al menos un resultado en las columnas de muestra."""
    for col in range(4, 4 + n_samples):
        if len(row) > col and row[col] is not None and str(row[col]).strip() != '':
            return True
    return False


def _set_samples_en_proceso(totals):
    """Regresa a 'En Proceso' las muestras (SamplingProcess) de los campos guardados.

    Cuando se registran resultados de metales pesados, la muestra pasa de
    'Confirmada' (o cualquier otro estado) a 'En Proceso' para indicar que
    está siendo analizada mediante el resultado masivo.
    """
    sample_ids = {entry['sample'].id for entry in totals.values()}
    if not sample_ids:
        return
    SamplingProcess.objects.filter(id__in=sample_ids).update(status_sampling='En Proceso')


def _resolve_sample_product(sample):
    """Retorna el producto asociado a la muestra (por punto o grupo de muestreo)."""
    if sample.point_sampling_id:
        return sample.point_sampling.product
    if sample.group_sampling_id:
        return sample.group_sampling.sampling_point.product
    return None


def _evaluate_comply(spec, total):
    """Evalúa el concepto: 'Cumple' si el total está dentro de los límites (inclusive).

    Retorna None cuando no hay especificación o sus límites son nulos.
    """
    if spec is None or (spec.lower_limit_prod is None and spec.upper_limit_prod is None):
        return None
    if spec.lower_limit_prod is not None and total < spec.lower_limit_prod:
        return 'No Cumple'
    if spec.upper_limit_prod is not None and total > spec.upper_limit_prod:
        return 'No Cumple'
    return 'Cumple'


def _update_sampling_analysis_results(totals, user):
    """Asigna la suma de resultados y el concepto al SamplingAnalysis de cada muestra+método.

    El average_concentration se redondea a las cifras significativas definidas
    en el método analítico (sig_figs_result) y el concepto se evalúa con ese valor.
    """
    products = {}
    for entry in totals.values():
        entry['product'] = _resolve_sample_product(entry['sample'])
        if entry['product']:
            products[entry['product'].id] = entry['product']

    specs = SpecificationProduct.objects.filter(
        product_id__in=products,
        method_test__analytical_method_id__in={e['method'].id for e in totals.values()},
    ).select_related('method_test')
    spec_map = {
        (spec.product_id, spec.method_test.analytical_method_id): spec for spec in specs
    }

    for entry in totals.values():
        analysis = SamplingAnalysis.objects.filter(
            sampling_process=entry['sample'],
            analytical_method=entry['method'],
        ).first()
        if analysis is None:
            analysis = SamplingAnalysis(
                sampling_process=entry['sample'],
                analytical_method=entry['method'],
                user_creation=user,
            )
        sig_figs = entry['method'].sig_figs_result or 2
        analysis.average_concentration = round(entry['total'], sig_figs)
        product_id = entry['product'].id if entry['product'] else None
        analysis.comply = _evaluate_comply(
            spec_map.get((product_id, entry['method'].id)), analysis.average_concentration
        )
        analysis.save()


def _get_result_unit(analysis, sampling_point=None):
    """Obtiene la unidad de medida del resultado del análisis.

    Precedencia: relación de cálculo, cálculos del método y, por último,
    la unidad definida en la SpecificationProduct del punto de muestreo.
    """
    relation = analysis.analytical_method_relation
    if relation and relation.unit_measure_calculate:
        return relation.unit_measure_calculate

    method = analysis.analytical_method
    if not method:
        return ''

    calc = method.analyticalmethodcalculate_set.first()
    if calc and calc.unit_measure_calculate:
        return calc.unit_measure_calculate

    calc_rel = method.analyticalmethodcalculaterelation_set.first()
    if calc_rel and calc_rel.unit_measure_calculate:
        return calc_rel.unit_measure_calculate

    from core.analytical_method.models import AnalyticalMethodCalculate
    fallback = AnalyticalMethodCalculate.objects.filter(
        analytical_method=method
    ).exclude(unit_measure_calculate__isnull=True).exclude(unit_measure_calculate="").first()
    if fallback:
        return fallback.unit_measure_calculate

    if sampling_point:
        specification = sampling_point.specification.filter(
            method_test__analytical_method=method
        ).exclude(unit_measure__isnull=True).exclude(unit_measure="").first()
        if specification:
            return specification.unit_measure

    return ''


def send_oss_notification_email(analysis):
    """Envía notificación por correo cuando un análisis resulta Fuera de Especificación (OSS).

    El correo se envía únicamente a los usuarios activos (is_active=True) que tengan
    marcada la opción de notificación de resultados OOS (notification_email_oss=True).
    Incluye el Punto de Muestreo, el Producto y la Planta, y un botón con acceso al
    listado de resultados fuera de especificación.
    """
    from django.conf import settings
    from django.core.mail import EmailMultiAlternatives
    from django.template.loader import render_to_string
    from django.urls import reverse

    if analysis.comply != 'No Cumple':
        return

    sampling_process = analysis.sampling_process
    sampling_point = (
        sampling_process.group_sampling.sampling_point
        if sampling_process.group_sampling
        else sampling_process.point_sampling
    )
    if not sampling_point:
        return

    product = sampling_point.product
    method = analysis.analytical_method
    method_desc = method.description_analytical_method if method else None

    subject = (
        f"Resultado (OOS) - {sampling_process.number_sample}"
        f" - {method_desc or ''}"
    )

    result_unit = _get_result_unit(analysis, sampling_point)
    result_display = (
        f"{analysis.average_concentration} {result_unit}".strip()
        if analysis.average_concentration is not None
        else "-"
    )

    list_path = reverse('sampling:list_sampling_process_out_specification')
    result_url = settings.SITE_URL.rstrip('/') + list_path

    html_content = render_to_string('emails/oss_notification.html', {
        'sampling_point': sampling_point,
        'product': product,
        'plant': product.site if product else None,
        'sample': sampling_process,
        'method': method,
        'analysis': analysis,
        'result_display': result_display,
        'result_url': result_url,
    })

    recipients = list(
        User.objects.filter(
            is_active=True,
            notification_email_oss=True,
        ).exclude(email='').values_list('email', flat=True)
    )
    if not recipients:
        return

    message = EmailMultiAlternatives(
        subject=subject,
        body='',
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=recipients,
    )
    message.attach_alternative(html_content, 'text/html')
    message.send()


FIXED_TEMPLATE_HEADERS = ('Metal', 'Metodo', 'Realizado por', 'Fecha')
SAMPLE_TEMPLATE_COLUMNS = 12


def build_massive_analysis_template(user):
    """Construye la plantilla Excel para el cargue masivo de metales pesados.

    La hoja "Datos" tiene las columnas fijas (Metal, Metodo, Realizado por,
    Fecha) y una columna por muestra desde la columna E en adelante; el usuario
    escribe el número de muestra (sampling_process) en la fila 1 y los
    resultados de cada metal en la columna de su muestra. La columna Metal
    viene pre-diligenciada con una fila por cada metal de los métodos
    habilitados del laboratorio del usuario.
    La hoja "Instrucciones" documenta el formato de cargue.

    Retorna un BytesIO con el contenido del archivo .xlsx.
    """
    workbook = Workbook()

    if user.laboratory:
        methods = AnalyticalMethod.objects.filter(
            laboratory=user.laboratory, enable_analytical_method=True,
            heavymetal__isnull=False,
        ).distinct().order_by('description_analytical_method')
    else:
        methods = AnalyticalMethod.objects.none()

    metals_by_method = {
        method: list(method.heavymetal_set.order_by('metal_description'))
        for method in methods.prefetch_related('heavymetal_set')
    }
    metal_headers = sorted({
        metal.metal_description
        for metals in metals_by_method.values()
        for metal in metals
    })

    # Hoja de datos: encabezados fijos + columna Metal pre-diligenciada por fila
    sheet = workbook.active
    sheet.title = 'Datos'
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', start_color='4472C4')
    sample_fill = PatternFill(fill_type='solid', start_color='EDEDED')
    headers = list(FIXED_TEMPLATE_HEADERS)
    widths = {'Metal': 22, 'Metodo': 22, 'Realizado por': 22, 'Fecha': 18}
    for index, title in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=index, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        sheet.column_dimensions[get_column_letter(index)].width = widths.get(title, max(len(title) + 4, 18))

    # Columnas de muestra desde la columna E (índice 5): vacías y estilizadas,
    # listas para escribir el número de muestra en la fila 1.
    for offset in range(SAMPLE_TEMPLATE_COLUMNS):
        column = 5 + offset
        cell = sheet.cell(row=1, column=column)
        cell.font = header_font
        cell.fill = sample_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        sheet.column_dimensions[get_column_letter(column)].width = 16

    # Una fila por cada metal disponible
    for offset, metal in enumerate(metal_headers):
        sheet.cell(row=2 + offset, column=1, value=metal)
    sheet.freeze_panes = 'E2'

    # Hoja de instrucciones
    info = workbook.create_sheet('Instrucciones')
    info.column_dimensions['A'].width = 110
    lines = [
        'Plantilla para el cargue masivo de análisis de Metales Pesados.',
        '',
        '1. Escriba el número de cada muestra (N° de Muestreo) en la fila 1 a partir de la columna E.',
        '2. Cada fila desde la 2 es el análisis de un metal (la columna Metal ya viene pre-diligenciada).',
        '3. Digite el resultado de cada metal en la columna de la muestra correspondiente.',
        '4. Deje vacía la celda si ese metal no aplica para esa muestra.',
        '5. Metal: descripción del metal. No modifique la columna A.',
        '6. Metodo: código o descripción del método analítico (ej: MET-01).',
        '7. Realizado por: usuario o nombre completo del analista registrado en el sistema.',
        '8. Fecha: formato aaaa-mm-dd o dd/mm/aaaa (puede incluir hora).',
        '9. Resultados de 0 o negativos se reemplazan por el límite de cuantificación del metal; '
        'si el metal no tiene límite configurado, el dato se reporta como novedad y no se carga.',
        '',
        'Metales por método de análisis:',
    ]
    if metals_by_method:
        for method, metals in metals_by_method.items():
            lines.append(
                f'- {method.code_analytical_method} {method.description_analytical_method}: '
                + ', '.join(metal.metal_description for metal in metals)
            )
    else:
        lines.append('- No hay métodos de análisis con metales configurados en su laboratorio.')
    for row, text in enumerate(lines, start=1):
        info.cell(row=row, column=1, value=text)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer
