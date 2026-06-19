from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Q, Avg
from django.http import JsonResponse, HttpResponse
from django.urls import reverse_lazy
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import ListView, TemplateView, View
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime
from django.utils import timezone

from core.mixins import ValidatePermissionRequiredMixin
from core.product.models import Product, SamplePoint, AnalyticalMethodProduct, SpecificationProduct
from core.sampling.models import SamplingAnalysis


class SamplingAnalysisListView(LoginRequiredMixin, ValidatePermissionRequiredMixin, ListView):
    model = SamplingAnalysis
    template_name = 'report/list_sampling_analysis.html'
    permission_required = 'reagent.add_reagent'

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST.get('action')
            if action == 'searchdata':
                product_id = request.POST.get('product')
                method_id = request.POST.get('analytical_method')

                if not method_id:
                    data = {
                        'columns': [],
                        'data': []
                    }
                else:
                    filters = Q(analytical_method_id=method_id)
                    if product_id:
                        filters &= (
                            Q(sampling_process__point_sampling__product_id=product_id) |
                            Q(sampling_process__group_sampling__sampling_point__product_id=product_id)
                        )

                    analyses = SamplingAnalysis.objects.filter(filters).select_related(
                        'sampling_process',
                        'sampling_process__point_sampling',
                        'sampling_process__group_sampling__sampling_point'
                    ).order_by('sampling_process__date_sampling')

                    # Obtener todos los puntos de muestreo asociados al producto
                    sample_points_query = SamplePoint.objects.filter(product_id=product_id).order_by('sequence')
                    sample_points = [p.sample_point_name for p in sample_points_query]

                    # Si no hay puntos de muestreo explícitos para el producto, al menos mostrar los que tienen datos
                    if not sample_points:
                        # Extraer nombres de puntos de muestreo de los análisis
                        # Consideramos tanto point_sampling como group_sampling.sampling_point
                        found_points = set()
                        for a in analyses:
                            if a.sampling_process.point_sampling:
                                found_points.add(a.sampling_process.point_sampling.sample_point_name)
                            elif a.sampling_process.group_sampling and a.sampling_process.group_sampling.sampling_point:
                                found_points.add(a.sampling_process.group_sampling.sampling_point.sample_point_name)
                        sample_points = sorted(list(found_points))

                    # Agrupar datos por fecha y hora
                    rows = {}
                    for a in analyses:
                        dt_str = a.sampling_process.date_sampling.strftime('%Y-%m-%d %H:%M:%S') if a.sampling_process.date_sampling else 'N/A'
                        if dt_str not in rows:
                            rows[dt_str] = {'date_analysis': dt_str}
                            # Inicializar todos los puntos conocidos con vacío o N/A para esta fila
                            for sp in sample_points:
                                rows[dt_str][sp] = '-'
                        
                        # Determinar el nombre del punto de muestreo del análisis actual
                        point_name = None
                        if a.sampling_process.point_sampling:
                            point_name = a.sampling_process.point_sampling.sample_point_name
                        elif a.sampling_process.group_sampling and a.sampling_process.group_sampling.sampling_point:
                            point_name = a.sampling_process.group_sampling.sampling_point.sample_point_name
                        
                        if point_name:
                            # Si el punto no estaba en la lista inicial (ej. producto mal configurado), lo agregamos
                            if point_name not in sample_points:
                                sample_points.append(point_name)
                                # Actualizar filas previas con este nuevo punto
                                for r_key in rows:
                                    if point_name not in rows[r_key]:
                                        rows[r_key][point_name] = '-'
                            
                            rows[dt_str][point_name] = a.average_concentration

                    data = {
                        'columns': sample_points,
                        'data': list(rows.values())
                    }
            elif action == 'search_analytical_method':
                data = []
                product_id = request.POST.get('id')
                if product_id:
                    from core.product.models import AnalyticalMethodProduct
                    methods = AnalyticalMethodProduct.objects.filter(product_id=product_id).select_related('analytical_method')
                    for m in methods:
                        data.append({
                            'id': m.analytical_method.id,
                            'text': f"{m.analytical_method.description_analytical_method}"
                        })
            elif action == 'search_sample_point':
                data = []
                product_id = request.POST.get('id')
                if product_id:
                    points = SamplePoint.objects.filter(product_id=product_id).order_by('sequence')
                    for p in points:
                        data.append({
                            'id': p.id,
                            'text': p.sample_point_name
                        })
            else:
                data['error'] = 'Ha ocurrido un error'
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data, safe=False)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Reporte de Análisis por Método'
        context['entity'] = 'Reporte de Análisis por Método de Análisis'
        context['div'] = '12'
        context['products'] = Product.objects.filter(enable_product=True)
        return context


class SamplingAnalysisChartView(LoginRequiredMixin, ValidatePermissionRequiredMixin, TemplateView):
    template_name = 'report/chart_sampling_analysis.html'
    permission_required = 'reagent.add_reagent'

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST.get('action')
            if action == 'get_graph_data':
                product_id = request.POST.get('product')
                method_id = request.POST.get('analytical_method')
                sample_point_id = request.POST.get('sample_point')
                date_from = request.POST.get('date_from')
                date_to = request.POST.get('date_to')

                if not product_id or not method_id:
                    data = {'categories': [], 'series': [], 'specifications': []}
                else:
                    filters = Q(analytical_method_id=method_id)
                    filters &= (
                        Q(sampling_process__point_sampling__product_id=product_id) |
                        Q(sampling_process__group_sampling__sampling_point__product_id=product_id)
                    )

                    if sample_point_id and sample_point_id != 'all':
                        filters &= (
                            Q(sampling_process__point_sampling_id=sample_point_id) |
                            Q(sampling_process__group_sampling__sampling_point_id=sample_point_id)
                        )

                    if date_from:
                        date_from_dt = timezone.make_aware(datetime.strptime(date_from, '%Y-%m-%d'))
                        filters &= Q(sampling_process__date_sampling__gte=date_from_dt)
                    if date_to:
                        date_to_dt = timezone.make_aware(datetime.strptime(f"{date_to} 23:59:59", '%Y-%m-%d %H:%M:%S'))
                        filters &= Q(sampling_process__date_sampling__lte=date_to_dt)

                    analyses = SamplingAnalysis.objects.filter(filters).select_related(
                        'sampling_process'
                    ).order_by('sampling_process__date_sampling')

                    # Especificaciones asociadas
                    specifications = []
                    if sample_point_id and sample_point_id != 'all':
                        sp = SamplePoint.objects.filter(id=sample_point_id).first()
                        if sp:
                            specs = sp.specification.filter(method_test__analytical_method_id=method_id)
                            for s in specs:
                                specifications.append({
                                    'name': s.test_prod,
                                    'lower_limit': s.lower_limit_prod,
                                    'upper_limit': s.upper_limit_prod,
                                    'sample_point': sp.sample_point_name,
                                    'unit_measure': s.unit_measure
                                })
                    else:
                        # Para 'all', obtener especificaciones de todos los puntos de muestreo del producto para ese método
                        points = SamplePoint.objects.filter(
                            product_id=product_id,
                            specification__method_test__analytical_method_id=method_id
                        ).prefetch_related('specification')

                        seen_specs = set()
                        for p in points:
                            p_specs = p.specification.filter(method_test__analytical_method_id=method_id)
                            for s in p_specs:
                                spec_key = (s.test_prod, s.lower_limit_prod, s.upper_limit_prod)
                                if spec_key not in seen_specs:
                                    specifications.append({
                                        'name': s.test_prod,
                                        'lower_limit': s.lower_limit_prod,
                                        'upper_limit': s.upper_limit_prod,
                                        'sample_point': p.sample_point_name,
                                        'unit_measure': s.unit_measure
                                    })
                                    seen_specs.add(spec_key)

                    data['specifications'] = specifications

                    # Agrupar por fecha y punto si es 'all', o solo por fecha si es uno específico
                    categories = []

                    if sample_point_id == 'all':
                        # Mostrar múltiples series, una por cada punto de muestreo
                        series_dict = {}
                        # Obtener todos los puntos de muestreo del producto para inicializar las series
                        points = SamplePoint.objects.filter(
                            product_id=product_id,
                            specification__method_test__analytical_method_id=method_id
                        ).distinct().order_by('sequence')
                        for p in points:
                            # Obtener unit_measure de la primera especificación del punto
                            unit_measure = ''
                            spec = p.specification.filter(method_test__analytical_method_id=method_id).first()
                            if spec and spec.unit_measure:
                                unit_measure = spec.unit_measure

                            series_dict[str(p.id)] = {
                                'name': p.sample_point_name,
                                'data': [],
                                'unit_measure': unit_measure  # ← AGREGADO
                            }

                        # Fechas únicas ordenadas
                        dates_raw = sorted(list(set(
                            a.sampling_process.date_sampling for a in analyses if a.sampling_process.date_sampling
                        )))
                        dates = [d.strftime('%Y-%m-%d %H:%M') for d in dates_raw]

                        data['categories'] = dates

                        for d_raw in dates_raw:
                            d_str = d_raw.strftime('%Y-%m-%d %H:%M')
                            # Filtrar análisis de este momento específico
                            moment_analyses = [a for a in analyses if a.sampling_process.date_sampling.strftime('%Y-%m-%d %H:%M') == d_str]

                            for p_id in series_dict:
                                # Buscar el valor para este día y punto
                                val = None
                                for a in moment_analyses:
                                    a_point_id = None
                                    if a.sampling_process.point_sampling_id:
                                        a_point_id = str(a.sampling_process.point_sampling_id)
                                    elif a.sampling_process.group_sampling and a.sampling_process.group_sampling.sampling_point_id:
                                        a_point_id = str(a.sampling_process.group_sampling.sampling_point_id)

                                    if a_point_id == p_id:
                                        val = a.average_concentration
                                        break
                                series_dict[p_id]['data'].append(val)

                        data['series'] = list(series_dict.values())
                    else:
                        # Una sola serie
                        point_name = 'Resultado'
                        unit_measure = ''
                        if sample_point_id:
                            sp = SamplePoint.objects.filter(id=sample_point_id).first()
                            if sp:
                                point_name = sp.sample_point_name
                                # Obtener unit_measure de las especificaciones del punto
                                spec = sp.specification.filter(method_test__analytical_method_id=method_id).first()
                                if spec and spec.unit_measure:
                                    unit_measure = spec.unit_measure

                        series_data = []
                        for a in analyses:
                            if a.sampling_process.date_sampling:
                                categories.append(a.sampling_process.date_sampling.strftime('%Y-%m-%d %H:%M'))
                                series_data.append(a.average_concentration)

                        data['categories'] = categories
                        data['series'] = [{
                            'name': point_name,
                            'data': series_data,
                            'unit_measure': unit_measure
                        }]

            elif action == 'search_analytical_method':
                data = []
                product_id = request.POST.get('id')

                if product_id:
                    methods = AnalyticalMethodProduct.objects.filter(product_id=product_id).select_related('analytical_method')

                    for m in methods:
                        data.append({
                            'id': str(m.analytical_method.id),
                            'text': f"{m.analytical_method.description_analytical_method}"
                        })
            elif action == 'search_sample_point':
                data = [{'id': 'all', 'text': 'Todos'}]
                product_id = request.POST.get('id')

                if product_id:
                    points = SamplePoint.objects.filter(product_id=product_id).order_by('sequence')

                    for p in points:
                        data.append({
                            'id': str(p.id),
                            'text': p.sample_point_name
                        })
            else:
                data['error'] = 'Ha ocurrido un error'
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data, safe=False)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Análisis Diario - Gráfico'
        context['entity'] = 'Gráfico de Análisis de Diario'
        # Aseguramos que los productos se ordenan por descripción
        context['products'] = Product.objects.filter(enable_product=True).order_by('description_product')
        context['icon'] = 'fa-solid fa-chart-line'
        context['list_url'] = reverse_lazy('report:sampling_analysis_chart') # URL de retorno
        return context


class SamplingAnalysisByPointListView(LoginRequiredMixin, ValidatePermissionRequiredMixin, ListView):
    model = SamplingAnalysis
    template_name = 'report/list_sampling_analysis_by_point.html'
    permission_required = 'reagent.add_reagent'

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST.get('action')
            if action == 'searchdata':
                product_id = request.POST.get('product')
                sample_point_id = request.POST.get('sample_point')

                if not sample_point_id:
                    data = {
                        'columns': [],
                        'data': []
                    }
                else:
                    filters = (
                        Q(sampling_process__point_sampling_id=sample_point_id) |
                        Q(sampling_process__group_sampling__sampling_point_id=sample_point_id)
                    )
                    
                    analyses = SamplingAnalysis.objects.filter(filters).select_related(
                        'sampling_process',
                        'analytical_method'
                    ).order_by('sampling_process__date_sampling')

                    # Obtener métodos analíticos asociados al producto
                    from core.product.models import AnalyticalMethodProduct
                    methods_query = AnalyticalMethodProduct.objects.filter(product_id=product_id).select_related('analytical_method')
                    methods = [m.analytical_method.description_analytical_method for m in methods_query]

                    # Si no hay métodos explícitos, usar los encontrados en los análisis
                    if not methods:
                        found_methods = set()
                        for a in analyses:
                            found_methods.add(a.analytical_method.description_analytical_method)
                        methods = sorted(list(found_methods))

                    # Agrupar datos por fecha y hora
                    rows = {}
                    for a in analyses:
                        dt_str = a.sampling_process.date_sampling.strftime('%Y-%m-%d %H:%M:%S') if a.sampling_process.date_sampling else 'N/A'
                        if dt_str not in rows:
                            rows[dt_str] = {'date_analysis': dt_str}
                            for m in methods:
                                rows[dt_str][m] = '-'
                        
                        method_name = a.analytical_method.description_analytical_method
                        if method_name not in methods:
                            methods.append(method_name)
                            for r_key in rows:
                                if method_name not in rows[r_key]:
                                    rows[r_key][method_name] = '-'
                        
                        rows[dt_str][method_name] = a.average_concentration

                    data = {
                        'columns': methods,
                        'data': list(rows.values())
                    }
            elif action == 'search_sample_point':
                data = []
                product_id = request.POST.get('id')
                if product_id:
                    points = SamplePoint.objects.filter(product_id=product_id).order_by('sequence')
                    for p in points:
                        data.append({
                            'id': p.id,
                            'text': p.sample_point_name
                        })
            else:
                data['error'] = 'Ha ocurrido un error'
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data, safe=False)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Reporte de Análisis por Punto de Muestreo'
        context['entity'] = 'Reporte de Análisis por Punto de Muestreo'
        context['div'] = '12'
        context['products'] = Product.objects.filter(enable_product=True)
        return context


class SamplingAnalysisByPointExcelView(LoginRequiredMixin, ValidatePermissionRequiredMixin, View):
    permission_required = 'reagent.add_reagent'

    def get(self, request, *args, **kwargs):
        try:
            product_id = request.GET.get('product')
            sample_point_id = request.GET.get('sample_point')

            if not sample_point_id:
                return HttpResponse("Debe seleccionar un punto de muestreo", status=400)

            filters = (
                Q(sampling_process__point_sampling_id=sample_point_id) |
                Q(sampling_process__group_sampling__sampling_point_id=sample_point_id)
            )

            analyses = SamplingAnalysis.objects.filter(filters).select_related(
                'sampling_process',
                'analytical_method'
            ).order_by('sampling_process__date_sampling')

            # Obtener métodos analíticos asociados al producto
            methods_query = AnalyticalMethodProduct.objects.filter(product_id=product_id).select_related(
                'analytical_method')
            methods = [m.analytical_method.description_analytical_method for m in methods_query]

            # Si no hay métodos explícitos, usar los encontrados en los análisis
            if not methods:
                found_methods = set()
                for a in analyses:
                    found_methods.add(a.analytical_method.description_analytical_method)
                methods = sorted(list(found_methods))

            # Agrupar datos por fecha y hora
            rows = {}
            for a in analyses:
                dt_str = a.sampling_process.date_sampling.strftime(
                    '%Y-%m-%d %H:%M:%S') if a.sampling_process.date_sampling else 'N/A'
                if dt_str not in rows:
                    rows[dt_str] = {'date_analysis': dt_str}
                    for m in methods:
                        rows[dt_str][m] = '-'

                method_name = a.analytical_method.description_analytical_method
                if method_name not in methods:
                    methods.append(method_name)
                    for r_key in rows:
                        if method_name not in rows[r_key]:
                            rows[r_key][method_name] = '-'

                rows[dt_str][method_name] = a.average_concentration

            # Crear el Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Reporte de Análisis"

            # Información de cabecera en el Excel
            product = Product.objects.get(pk=product_id)
            sample_point = SamplePoint.objects.get(pk=sample_point_id)
            
            ws.merge_cells('A1:C1')
            ws['A1'] = f"Producto: {product.description_product}"
            ws.merge_cells('A2:C2')
            ws['A2'] = f"Punto de Muestreo: {sample_point.sample_point_name}"
            ws['A1'].font = Font(bold=True, size=12)
            ws['A2'].font = Font(bold=True, size=12)

            # Estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="2A383E", end_color="2A383E", fill_type="solid")
            alignment = Alignment(horizontal="center")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Cabeceras
            headers = ['Fecha y Hora'] + methods
            header_row = 4
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment
                cell.border = border

            # Datos
            for row_num, (dt_str, row_data) in enumerate(rows.items(), header_row + 1):
                ws.cell(row=row_num, column=1, value=dt_str).border = border
                for col_num, method in enumerate(methods, 2):
                    ws.cell(row=row_num, column=col_num, value=row_data.get(method, '-')).border = border

            # Ajustar ancho de columnas
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="reporte_analisis_{datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx"'
            wb.save(response)
            return response

        except Exception as e:
            return HttpResponse(f"Error al generar el excel: {str(e)}", status=500)
