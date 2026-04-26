from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import JsonResponse, HttpResponse, HttpResponseRedirect
from django.urls import reverse_lazy, reverse
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import ListView, CreateView, UpdateView, View, DetailView

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime, timedelta
from django.utils import timezone
from django.utils.safestring import mark_safe
from core.condition.forms import ConditionForm, ConditionRegisterForm, ConditionRegisterActionsForm
from core.condition.models import Condition, ConditionRegister
from core.mixins import ValidatePermissionRequiredMixin


class ConditionVariableAPI(LoginRequiredMixin, View):
    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST.get('action')
            if action == 'get_variable':
                condition_id = request.POST.get('id')
                if condition_id:
                    condition = Condition.objects.get(pk=condition_id)
                    data['variable'] = condition.variable
                    data['upper_limit'] = condition.upper_limit
                    data['lower_limit'] = condition.lower_limit
                else:
                    data['error'] = 'No se ha proporcionado el ID de la condición'
            else:
                data['error'] = 'No ha ingresado una acción'
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data)


class ConditionListView(LoginRequiredMixin, ValidatePermissionRequiredMixin, ListView):
    model = Condition
    template_name = 'condition/list_condition.html'
    permission_required = 'condition.view_condition'

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST['action']
            if action == 'searchdata':
                data = []
                for i in Condition.objects.all():
                    data.append({
                        'id': i.id,
                        'laboratory': i.laboratory.laboratory_name if i.laboratory else 'N/A',
                        'area': i.area,
                        'variable': i.variable,
                        'upper_limit': i.upper_limit,
                        'lower_limit': i.lower_limit,
                        'enabled': 'Sí' if i.enabled else 'No',
                    })
            elif action == 'search_graph':
                data = []
                condition_id = request.POST['id']
                condition = Condition.objects.get(pk=condition_id)
                registers = ConditionRegister.objects.filter(condition_id=condition_id).order_by('-registration_date')[:20]
                for i in reversed(registers):
                    data.append({
                        'date': timezone.localtime(i.registration_date).strftime('%Y-%m-%d %H:%M'),
                        'data': i.registered_data,
                        'upper_limit': condition.upper_limit,
                        'lower_limit': condition.lower_limit,
                    })
            else:
                data['error'] = 'Ha ocurrido un error'
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data, safe=False)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Áreas y Condiciones'
        context['create_url'] = reverse_lazy('condition:create_condition')
        context['list_url'] = reverse_lazy('condition:list_condition')
        context['entity'] = 'Áreas y Condiciones'
        context['div'] = '12'
        return context


class ConditionCreateView(LoginRequiredMixin, ValidatePermissionRequiredMixin, CreateView):
    model = Condition
    form_class = ConditionForm
    template_name = 'condition/create_condition.html'
    permission_required = 'condition.add_condition'

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST['action']
            if action == 'add':
                form = self.get_form()
                if form.is_valid():
                    form.save()
                    messages.success(request, '¡Condición registrada satisfactoriamente!')
                    data['success'] = True
                    data['redirect_url'] = self.get_success_url()
                else:
                    error_list = []
                    for field, errors in form.errors.items():
                        for error in errors:
                            if field == '__all__':
                                error_list.append(f"{error}")
                            else:
                                error_list.append(f"<b>{form.fields[field].label}</b>: {error}")
                    data['error'] = "<br>".join(error_list)
            else:
                data['error'] = 'No ha ingresado datos en los campos'
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data)

    def get_success_url(self):
        return reverse('condition:list_condition')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Creación de Área y Condiciones Ambientales'
        context['entity'] = 'Creación de Área y Condiciones Ambientales'
        context['list_url'] = reverse_lazy('condition:list_condition')
        context['action'] = 'add'
        return context


class ConditionUpdateView(LoginRequiredMixin, ValidatePermissionRequiredMixin, UpdateView):
    model = Condition
    form_class = ConditionForm
    template_name = 'condition/create_condition.html'
    permission_required = 'condition.change_condition'

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST['action']
            if action == 'edit':
                form = self.get_form()
                if form.is_valid():
                    form.save()
                    messages.success(request, '¡Condición actualizada satisfactoriamente!')
                    data['success'] = True
                    data['redirect_url'] = self.get_success_url()
                else:
                    error_list = []
                    for field, errors in form.errors.items():
                        for error in errors:
                            if field == '__all__':
                                error_list.append(f"{error}")
                            else:
                                error_list.append(f"<b>{form.fields[field].label}</b>: {error}")
                    data['error'] = "<br>".join(error_list)
            else:
                data['error'] = 'No ha ingresado datos en los campos'
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data)

    def get_success_url(self):
        return reverse('condition:list_condition')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Edición de Área y Condición Ambiental'
        context['entity'] = 'Edición de Área y Condición Ambiental'
        context['list_url'] = reverse_lazy('condition:list_condition')
        context['action'] = 'edit'
        return context


class ConditionRegisterListView(LoginRequiredMixin, ValidatePermissionRequiredMixin, ListView):
    model = ConditionRegister
    template_name = 'condition/list_condition_register.html'
    permission_required = 'condition.view_conditionregister'

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST.get('action')
            if action == 'searchdata':
                registers = list(ConditionRegister.objects.select_related(
                    'condition',
                    'registered_by',
                    'actions_registered_by'
                ).values(
                    'id',
                    'registration_date',
                    'registered_by__first_name',
                    'registered_by__last_name',
                    'registered_data',
                    'condition__area',
                    'condition__variable',
                    'condition__upper_limit',
                    'condition__lower_limit',
                    'actions_registered_by__id'
                ).order_by('-registration_date'))

                for r in registers:
                    r['registration_date'] = timezone.localtime(r['registration_date']).strftime('%Y-%m-%d %H:%M:%S')
                    first_name = r.pop('registered_by__first_name') or ''
                    last_name = r.pop('registered_by__last_name') or ''
                    r['registered_by'] = f"{first_name} {last_name}".strip()
                    unit = '%' if r['condition__variable'] == 'Humedad Relativa' else '°C'
                    r['registered_data_formatted'] = f"{r['registered_data']}{unit}"
                    r['actions_registered_by'] = r.pop('actions_registered_by__id')

                return JsonResponse(registers, safe=False)
            else:
                data['error'] = 'Ha ocurrido un error'
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data, safe=False)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Registros de Condiciones Ambientales'
        context['create_url'] = reverse_lazy('condition:create_condition_register')
        context['list_url'] = reverse_lazy('condition:list_condition_register')
        context['export_url'] = reverse_lazy('condition:export_condition_register_excel')
        context['entity'] = 'Registros de Condiciones Ambientales'
        context['div'] = '12'
        return context


class ConditionRegisterExportExcelView(LoginRequiredMixin, ValidatePermissionRequiredMixin, View):
    permission_required = 'condition.view_conditionregister'

    def get(self, request, *args, **kwargs):
        try:
            # Filtrar registros de los últimos 3 años
            three_years_ago = timezone.now() - timedelta(days=3 * 365)
            registers = ConditionRegister.objects.filter(
                registration_date__gte=three_years_ago
            ).select_related('condition', 'registered_by', 'actions_registered_by').order_by('-registration_date')

            # Crear libro y hoja de Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Registros Condiciones"

            # Estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Encabezados
            headers = [
                'Fecha de Registro', 'Registrado por', 'Área', 'Variable',
                'Dato Registrado', 'Límite Inferior', 'Límite Superior',
                'Acciones', 'Acciones Registradas por'
            ]

            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment
                cell.border = border
                # Ajustar ancho de columna (aproximado)
                ws.column_dimensions[cell.column_letter].width = 20

            # Datos
            for row_num, r in enumerate(registers, 2):
                data = [
                    timezone.localtime(r.registration_date).strftime('%Y-%m-%d %H:%M:%S'),
                    r.registered_by.get_full_name(),
                    r.condition.area,
                    r.condition.variable,
                    r.registered_data,
                    r.condition.lower_limit,
                    r.condition.upper_limit,
                    r.actions or '',
                    r.actions_registered_by.get_full_name() if r.actions_registered_by else ''
                ]

                for col_num, value in enumerate(data, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.alignment = alignment
                    cell.border = border

            # Preparar respuesta
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = f'attachment; filename="Registros_Condiciones_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            
            wb.save(response)
            return response

        except Exception as e:
            messages.error(request, f'Error al generar el Excel: {str(e)}')
            return HttpResponseRedirect(reverse_lazy('condition:list_condition_register'))


class ConditionRegisterCreateView(LoginRequiredMixin, ValidatePermissionRequiredMixin, CreateView):
    model = ConditionRegister
    form_class = ConditionRegisterForm
    template_name = 'condition/create_condition_register.html'
    permission_required = 'condition.add_conditionregister'

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST['action']
            if action == 'add':
                form = ConditionRegisterForm(request.POST)
                if form.is_valid():
                    form.save()
                    messages.success(request, 'Condiciones ambientales registradas satisfactoriamente!')
                    data['success'] = True
                    data['redirect_url'] = self.get_success_url()
                else:
                    error_list = []
                    for field, errors in form.errors.items():
                        for error in errors:
                            if field == '__all__':
                                error_list.append(f"{error}")
                            else:
                                error_list.append(f"<b>{form.fields[field].label}</b>: {error}")
                    data['error'] = "<br>".join(error_list)
            else:
                data['error'] = 'No ha ingresado datos en los campos'
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data)

    def get_success_url(self):
        return reverse('condition:list_condition_register')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Registro de Condiciones Ambientales'
        context['entity'] = 'Registro de Condiciones Ambientales'
        context['list_url'] = reverse_lazy('condition:list_condition_register')
        context['action'] = 'add'
        return context


class ConditionRegisterUpdateView(LoginRequiredMixin, ValidatePermissionRequiredMixin, UpdateView):
    model = ConditionRegister
    form_class = ConditionRegisterForm
    template_name = 'condition/create_condition_register.html'
    permission_required = 'condition.change_conditionregister'

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST['action']
            if action == 'edit':
                form = self.get_form()
                if form.is_valid():
                    form.save()
                    messages.success(request, '¡Registro de condición actualizado satisfactoriamente!')
                    data['success'] = True
                    data['redirect_url'] = self.get_success_url()
                else:
                    error_list = []
                    for field, errors in form.errors.items():
                        for error in errors:
                            if field == '__all__':
                                error_list.append(f"{error}")
                            else:
                                error_list.append(f"<b>{form.fields[field].label}</b>: {error}")
                    data['error'] = "<br>".join(error_list)
            else:
                data['error'] = 'No ha ingresado datos en los campos'
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data)

    def get_success_url(self):
        return reverse('condition:list_condition_register')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Edición de Registro de Condiciones Ambientales'
        context['entity'] = 'Edición de Registro de Condiciones Ambientales   '
        context['list_url'] = reverse_lazy('condition:list_condition_register')
        context['action'] = 'edit'
        return context


class ConditionRegisterActionsUpdateView(LoginRequiredMixin, ValidatePermissionRequiredMixin, UpdateView):
    model = ConditionRegister
    form_class = ConditionRegisterActionsForm
    template_name = 'modal_two.html'
    permission_required = 'condition.change_conditionregister'

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST['action']
            if action == 'edit':
                form = self.get_form()
                if form.is_valid():
                    form.save()
                    messages.success(request, '¡Acciones registradas satisfactoriamente!')
                    data['success'] = True
                else:
                    error_list = []
                    for field, errors in form.errors.items():
                        for error in errors:
                            if field == '__all__':
                                error_list.append(f"{error}")
                            else:
                                error_list.append(f"<b>{form.fields[field].label}</b>: {error}")
                    data['error'] = "<br>".join(error_list)
            else:
                data['error'] = 'No ha ingresado datos en los campos'
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Registro de Acciones o Correcciones'
        context['entity'] = 'Registro de Acciones o Correcciones'
        
        registered_data = self.object.registered_data
        upper_limit = self.object.condition.upper_limit
        lower_limit = self.object.condition.lower_limit
        
        info_text = f"Condición: {self.object.condition} | Dato: {registered_data}"
        if registered_data > upper_limit or registered_data < lower_limit:
            info_text = mark_safe(f"Condición: {self.object.condition} | <span style='color: red;'>Dato: {registered_data} (Fuera de rango)</span>")
            
        context['info_form'] = info_text
        context['action'] = 'edit'
        return context


class ConditionRegisterDetailView(LoginRequiredMixin, ValidatePermissionRequiredMixin, DetailView):
    model = ConditionRegister
    template_name = 'condition/detail_condition_register.html'
    permission_required = 'condition.view_conditionregister'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Detalle de Acciones Registradas'
        context['entity'] = 'Detalle de Acciones Registradas'
        return context
